import openpyxl
import json
from datetime import datetime

FILENAME = "06_REGISTRO_VENTAS(VERSION FINAL).xlsm"

def serialize_value(val):
    if isinstance(val, (datetime,)):
        return val.isoformat()
    return val

def extract_data():
    try:
        # Cargamos con data_only=True para obtener los valores calculados
        wb = openpyxl.load_workbook(FILENAME, data_only=True)
        
        # Categorías predefinidas
        categorias_set = set(["Panadería", "Abarrotes", "Bebidas", "Confitería"])
        
        # Extraer Productos desde 'CODIGOS'
        productos = []
        if "CODIGOS" in wb.sheetnames:
            sheet = wb["CODIGOS"]
            
            # Primero, encontramos la fila de encabezados
            header_row = None
            for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=15), 1):
                first_cell = str(row[0].value).upper() if row[0].value else ""
                if "CODIGO" in first_cell or "BARR" in first_cell:
                    header_row = i
                    break
            
            if not header_row:
                print("⚠️  No se encontró la fila de encabezados. Usando fila 9 por defecto.")
                header_row = 9
            
            print(f"📋 Encabezados encontrados en fila {header_row}")
            print(f"📦 Extrayendo productos desde fila {header_row + 1}...")
            
            # Leemos los encabezados para mapear columnas
            headers = [str(cell.value).strip().upper() if cell.value else f"COL_{i}" 
                      for i, cell in enumerate(sheet[header_row])]
            
            print(f"Columnas detectadas: {headers[:10]}")
            
            # Intentamos identificar las columnas clave
            col_barcode = next((i for i, h in enumerate(headers) if "CODIGO" in h or "BARR" in h), 0)
            col_name = next((i for i, h in enumerate(headers) if "DETALLE" in h or "NOMBRE" in h or "PRODUCTO" in h), 1)
            col_stock = next((i for i, h in enumerate(headers) if "INVENTARIO" in h or "STOCK" in h or "CANTIDAD" in h), 2)
            col_cost = next((i for i, h in enumerate(headers) if "COSTO" in h or "COMPRA" in h), 4)
            col_price = next((i for i, h in enumerate(headers) if "VENTA" in h or "PRECIO" in h), 5)
            
            print(f"Mapeo de columnas:")
            print(f"  - Código de barras: columna {col_barcode} ({headers[col_barcode]})")
            print(f"  - Nombre: columna {col_name} ({headers[col_name]})")
            print(f"  - Stock: columna {col_stock} ({headers[col_stock]})")
            print(f"  - Costo: columna {col_cost} ({headers[col_cost]})")
            print(f"  - Precio: columna {col_price} ({headers[col_price]})")
            
            # Extraemos los datos
            for row in sheet.iter_rows(min_row=header_row + 1, max_col=15):
                barcode = serialize_value(row[col_barcode].value)
                name = serialize_value(row[col_name].value)
                stock = serialize_value(row[col_stock].value)
                cost = serialize_value(row[col_cost].value)
                price = serialize_value(row[col_price].value)
                
                # Solo procesamos si hay nombre
                if not name or str(name).strip() == "":
                    continue
                
                # Limpieza de código de barras
                if barcode and ("escanear" in str(barcode).lower() or barcode == ""):
                    barcode = None
                
                # Detectar si es pesable
                name_lower = str(name).lower()
                es_pesable = any(keyword in name_lower for keyword in ["kg", "granel", "hallulla", "marraqueta", "pan corriente"])
                
                # Los precios del Excel ya están en CLP, NO multiplicar
                precio_costo_clp = float(cost) if cost and cost != "" else 0
                precio_venta_clp = float(price) if price and price != "" else 0
                
                # Calcular margen correctamente: (venta - costo) / costo
                margen = 0.5  # Default
                if precio_costo_clp > 0 and precio_venta_clp > 0:
                    margen = (precio_venta_clp - precio_costo_clp) / precio_costo_clp
                
                productos.append({
                    "codigo_barras": str(barcode) if barcode else None,
                    "nombre": str(name).strip(),
                    "stock_cantidad": float(stock) if stock and stock != "" else 0,
                    "precio_costo": precio_costo_clp,
                    "precio_venta": precio_venta_clp,
                    "margen_porcentaje": round(margen, 4),
                    "es_pesable": es_pesable
                })
        
        # Consolidar resultado
        data = {
            "categorias": list(categorias_set),
            "productos": productos,
            "metadata": {
                "fecha_extraccion": datetime.now().isoformat(),
                "total_productos": len(productos),
                "archivo_origen": FILENAME
            }
        }
        
        with open("inventario_consolidado.json", "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        print(f"\n✅ Éxito: Se han extraído {len(productos)} productos")
        print(f"📄 Archivo guardado: inventario_consolidado.json")
        print(f"\n🔍 Primeros 3 productos extraídos:")
        for i, p in enumerate(productos[:3], 1):
            print(f"  {i}. {p['nombre']}")
            print(f"     - Código: {p['codigo_barras']}")
            print(f"     - Precio venta: ${p['precio_venta']}")
            print(f"     - Pesable: {'Sí' if p['es_pesable'] else 'No'}")
    
    except Exception as e:
        print(f"❌ Error durante la extracción: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    extract_data()
