import openpyxl
import json
from datetime import datetime

FILENAME = "06_REGISTRO_VENTAS(VERSION FINAL).xlsm"

def serialize_value(val):
    if isinstance(val, (datetime,)):
        return val.isoformat()
    return val

def extract_data():
    try:
        # Cargamos con data_only=True para obtener los valores calculados por las fórmulas
        wb = openpyxl.load_workbook(FILENAME, data_only=True)
        
        # 1. Extraer Categorías (Basado en el análisis previo de la hoja inicial)
        # Identificamos categorías comunes de los datos
        categorias_set = set(["Panadería", "Abarrotes", "Bebidas", "Confitería"])
        
        # 2. Extraer Productos desde 'CODIGOS'
        productos = []
        if "CODIGOS" in wb.sheetnames:
            sheet = wb["CODIGOS"]
            # Saltamos encabezados
            for row in sheet.iter_rows(min_row=9, max_col=13): # Basado en análisis previo, fila 9 empieza la data
                barcode = serialize_value(row[0].value)
                name = serialize_value(row[1].value)
                stock = serialize_value(row[2].value)
                cost = serialize_value(row[4].value) # Columna E es costo
                margin = 0.5 # Valor por defecto o extraído si es posible
                
                if barcode and name:
                    productos.append({
                        "codigo_barras": str(barcode),
                        "nombre": str(name),
                        "stock_cantidad": float(stock) if stock is not None else 0,
                        "precio_costo": float(cost) if cost is not None else 0,
                        "margen_porcentaje": margin,
                        "es_pesable": "kg" in str(name).lower() or "granel" in str(name).lower()
                    })

        # 3. Consolidar resultado
        data = {
            "categorias": list(categorias_set),
            "productos": productos,
            "metadata": {
                "fecha_extraccion": datetime.now().isoformat(),
                "total_productos": len(productos)
            }
        }

        with open("inventario_consolidado.json", "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        print(f"Éxito: Se han extraído {len(productos)} productos y se han guardado en inventario_consolidado.json")

    except Exception as e:
        print(f"Error durante la extracción: {e}")

if __name__ == "__main__":
    extract_data()
